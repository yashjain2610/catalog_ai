from openpyxl import load_workbook

def get_fixed_values_dict(file_path):

# Load the workbook and select the active sheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Assuming header is in the first row
    headers = {cell.value: idx for idx, cell in enumerate(ws[1])}

    # Get column indices
    attr_col = headers.get("Attributes")
    ai_col = headers.get("AI")
    fixed_col = headers.get("Fixed")
    value_col = headers.get("Value")

    # Result dictionary
    result = {}

    target_fields = []

    # Iterate through rows (skip header)
    for row in ws.iter_rows(min_row=2, values_only=True):
        ai = row[ai_col]
        fixed = row[fixed_col]
        
        if (str(ai).strip().lower() == "yes" or str(ai).strip().lower() == "ai") and str(fixed).strip().lower() == "yes":
            attr = row[attr_col]
            value = row[value_col]
            if attr and value:
                result[attr] = value
        elif (str(ai).strip().lower() == "yes" or str(ai).strip().lower() == "ai"):
            attr = row[attr_col]
            target_fields.append(attr)

    return result, target_fields



target_fields_earrings = [
            "Seller SKU ID", "Type", "Color", "Base Material", "Gemstone", "Pearl Type", "Collection",
            "Occasion", "Piercing Required", "Number of Gemstones", "Earring Back Type",
            "Finish", "Design", "Metal Color", "Natural/Synthetic Diamond", "Pearl Shape", "Pearl Grade", "Pearl Color",
            "Other Features", "Search Keywords", "Key Features", "Trend", "Closure Type", "Sub Type",
            "Earring Shape", "With Ear Chain", "Earring Set Type", "Ornamentation Type","Trend AW 16", "Description","Pearl Length (mm)","Pearl Diameter (mm)",
            "Width (mm)","Height (mm)","Diameter (mm)"
        ]

fixed_values_earrings = {
    'Listing Status': 'Active',
    'Fullfilment by': 'Seller',
    'Procurement type': 'express',
    'Procurement SLA (DAY)': '1',
    'Shipping provider': 'Flipkart',
    'Local delivery charge (INR)': '0',
    'Zonal delivery charge (INR)': '0',
    'National delivery charge (INR)': '16',
    'Weight (KG)': '0.3',
    'Height (CM)': '1.7',
    'Length (CM)': '2.4',
    'Breadth (CM)': '2.4',
    'HSN': '71171900',
    'Country Of Origin': 'India',
    'Manufacturer Details': 'Alya Jewels',
    'Packer Details': 'Alya Jewels',
    'Tax Code': 'GST_3',
    'Minimum Order Quantity (MinOQ)': '1',
    'Brand': 'Alya Jewels',
    'Ideal For': 'Girls: Women',
    'Plating': 'NA',
    'Setting': 'Affixed',
    'Weight (g)': '150',
    'Net Quantity': '1',
    'Diamond Clarity': 'NA'
}

fixed_values_bracelet_flipkart = {
    "Listing Status": "Active",
    "Fullfilment by": "Seller",
    "Procurement type": "express",
    "Procurement SLA (DAY)": "1",
    "Shipping provider": "Flipkart",
    "National delivery charge (INR)": "0",
    "Weight (KG)": "0.2",
    "HSN": "71171900",
    "Country Of Origin": "India",
    "Manufacturer Details": "Alya Jewels",
    "Packer Details": "Alya Jewels",
    "Tax Code": "GST_3",
    "Minimum Order Quantity (MinOQ)": "1",
    "Brand": "Alya Jewels",
    "Pack of (pieces)": "1",
    "Ideal For": "Girls: Women",
    "Plating": "NA",
    "Gemstone": "NA",
    "Diamond Clarity": "NA",
    "Diamond Color Grade": "NA",
    "Pearl Type": "NA",
    "Setting": "Affixed",
    "Weight (g)": "150",
    "Sales Package": "1"
}

target_fields_bracelet_flipkart = [
    "Seller SKU ID","Type", "Color", "Base Material", "Collection", "Occasion", "Body Structure", "Number of Gemstones", "Design",
    "Finish", "Tags/Charms Attachable", "Tag/Charm Shape", "Clasp", "Stretchable", "Adjustable Length",
    "Metal Color", "Natural/Synthetic Diamond", "Pearl Color", "Pearl Shape", "Pearl Grade", "Product Title",
    "Other Features", "Trend", "Search Keywords", "Key Features", "Style Type", "Ornamentation Type", "Brand Color","Description"
]

fixed_values_necklace_flipkart = {
    'Listing Status': 'Active',
    'Fullfilment by': 'Seller',
    'Procurement type': 'express',
    'Procurement SLA (DAY)': '1',
    'Shipping provider': 'Flipkart',
    'Local delivery charge (INR)': '0',
    'Zonal delivery charge (INR)': '0',
    'National delivery charge (INR)': '0',
    'HSN': '71171900',
    'Country Of Origin': 'India',
    'Manufacturer Details': 'Alya Jewels',
    'Packer Details': 'Alya Jewels',
    'Tax Code': 'GST_3',
    'Minimum Order Quantity (MinOQ)': '1',
    'Brand': 'alya jewels ',
    'Pack of': '1',
    'Sales Package': '1'
}

target_fields_necklace_flipkart =  [
    'Seller SKU ID',
    'Base Material',
    'Type',
    'Color',
    'Gemstone',
    'Ideal For',
    'Certification',
    'Plating',
    'Setting',
    'Occasion',
    'Diamond Color Grade',
    'Pearl Type',
    'Collection',
    'Diamond Cut',
    'Finish',
    'Number of Gemstones',
    'Necklace Type',
    'Clasp Type',
    'Pearl Color',
    'Pearl Shape',
    'Other Features',
    'Search Keywords',
    'Key Features',
    'Trend',
    'Necklace Style',
    'Necklace Length',
    'Work Finish',
    'Chain Style',
    'Ornamentation Type',
    'Pendant Shape',
    'Necklace Design',
    'Necklace Width',
    'Necklace Color',
    'Description'
]

fixed_values_necklace_amz = {
    'feed_product_type': 'necklace',
    'brand_name': 'Alya Jewels',
    'recommended_browse_nodes': '7124388031',
    'manufacturer': 'Alya Jewels',
    'target_gender': 'Female',
    'packer_contact_information': 'Alya Jewels',
    'manufacturer_contact_information': 'Alya Jewels',
    'gender': 'Women',
    'size_name': 'Free',
    'unit_count': '1',
    'supplier_declared_dg_hz_regulation1': 'Not Applicable',
    'country_of_origin': 'India',
    'external_product_information': '71171900',
    'metals_metal_stamp': 'No Metal Stamp',
    'currency': 'INR',
    'number_of_items': '1',
    'fulfillment_latency': '1',
    'offering_can_be_giftwrapped': 'No',
    'product_tax_code': 'A_GEN_JEWELLERY',
    'condition_type': 'New'
}

target_fields_necklace_amz = [
    'item_sku',
    'item_name',
    'stone_color',
    'style_name',
    'stones_color',
    'item_type_name',
    'occasion_type1',
    'clasp_type',
    'color_map',
    'collection_name',
    'material_type',
    'stones_type',
    'number_of_stones',
    'setting_type',
    'bullet_point1',
    'bullet_point2',
    'bullet_point3',
    'bullet_point4',
    'bullet_point5',
    'color_name',
    'stone_shape',
    'chain_length_unit',
    'theme',
    'stone_cut',
    'chain_type',
    'lifecycle_supply_type',
    'gem_type',
    'pearl_type',
    'pearl_minimum_color',
    'number_of_pearls',
    'product_description'
]




#fixed_values_earrings_amz , target_fields_earrings_amz = get_fixed_values_dict("Logic_of_earrings_amz.xlsx")

fixed_values_earrings_amz = {
    'item_type_name': 'earring',
    'update_delete': 'Update',
    'brand_name': 'Alya Jewels',
    'recommended_browse_nodes': '2152555031',
    'target_gender': 'Women',
    'packer_contact_information': 'Alya Jewels',
    'metals_metal_stamp': 'No Metal Stamo',
    'size_name': 'Free',
    'item_width_unit_of_measure': 'Cm',
    'item_height_unit_of_measure': 'Cm',
    'item_length_unit_of_measure': 'Cm',
    'supplier_declared_dg_hz_regulation1': 'Not Applicable',
    'country_of_origin': 'India',
    'external_product_information': '71171900',
    'currency': 'INR',
    'fulfillment_latency': '1',
    'offering_can_be_giftwrapped': 'NO',
    'product_tax_code': 'A_GEN_JEWELLERY',
    'offering_can_be_gift_messaged': 'NO',
    'condition_type': 'New'
}

# #target_fields_earrings_amz = [
#     'Seller SKU',
#     'Item Name (aka Title)', 'Stone Colour',
#     'Style', 'Stone Color', 'Item Type Name', 'Occasion','Clasp Type', 'Color Map',
#     'Material Type FREE', 'Stone Shape', 'Stone Clarity', 'Collection',
#     'Metal Type', 'Material Type', 'Number of Stones', 'Setting Type',
#     'Bullet Point', 'Bullet Point', 'Bullet Point', 'Bullet Point',
#     'Bullet Point', 'Bullet Point', 'Bullet Point', 'Bullet Point',
#     'Bullet Point', 'Bullet Point', 'Color', 'Stone Shape', 'Theme',
#     'Gender', 'Earring Back Finding', 'Gemstone Type'
# ]
#print(fixed_values_earrings_amz)

#{'feed_product_type': 1, 'item_sku': 2, 'update_delete': 3, 'brand_name': 4, 'external_product_id': 5, 'external_product_id_type': 6, 'item_name': 7, 'recommended_browse_nodes': 8, 'model': 9, 'product_description': 10, 'part_number': 11, 'grade_rating': 12, 'manufacturer': 13, 'certificate_type': 14, 'quantity': 15, 'standard_price': 16, 'age_range_description': 17, 'target_gender': 18, 'main_image_url': 19, 'other_image_url1': 20, 'other_image_url2': 21, 'other_image_url3': 22, 'other_image_url4': 23, 'other_image_url5': 24, 'other_image_url6': 25, 'other_image_url7': 26, 'other_image_url8': 27, 'swatch_image_url': 28, 'relationship_type': 29, 'variation_theme': 30, 'parent_sku': 31, 'parent_child': 32, 'stone_color': 33, 'style_name': 34, 'metals_metal_weight_unit_of_measure': 35, 'drop_length_unit': 36, 'stones_color': 37, 'metal_type': 38, 'packer_contact_information': 39, 'item_type_name': 40, 'occasion_type1': 41, 'occasion_type2': 42, 'occasion_type3': 43, 'occasion_type4': 44, 'occasion_type5': 45, 'certificate_number': 46, 'clasp_type': 47, 'stones_id': 48, 'stones_creation_method': 49, 'color_map': 50, 'material_type_free1': 51, 'material_type_free2': 52, 'material_type_free3': 53, 'material_type_free4': 54, 'material_type_free5': 55, 'subject_character': 56, 'stone_shape': 57, 'stone_clarity': 58, 'collection_name': 59, 'metals_metal_weight': 60, 'importer_contact_information': 61, 'metals_metal_type': 62, 'material_type': 63, 'stones_type': 64, 'stones_clarity': 65, 'stones_number_of_stones': 66, 'manufacturer_contact_information': 67, 'stones_treatment_method': 68, 'fur_description': 69, 'setting_type': 70, 'metals_id': 71, 'bullet_point1': 72, 'bullet_point2': 73, 'bullet_point3': 74, 'bullet_point4': 75, 'bullet_point5': 76, 'bullet_point6': 77, 'bullet_point7': 78, 'bullet_point8': 79, 'bullet_point9': 80, 'bullet_point10': 81, 'generic_keywords': 82, 'initial_character': 83, 'metals_metal_stamp': 84, 'color_name': 85, 'stones_shape': 86, 'ring_size': 87, 'chain_length_unit': 88, 'theme': 89, 'stones_cut': 90, 'stone_cut': 91, 'department_name': 92, 'back_finding': 93, 'drop_length': 94, 'size_name': 95, 'chain_length_derived': 96, 'chain_type': 97, 'lifecycle_supply_type': 98, 'gem_type': 99, 'item_booking_date': 100, 'pearl_type': 101, 'material_features': 102, 'pearl_uniformity': 103, 'pearl_shape': 104, 'pearl_surface_blemishes': 105, 'pearl_lustre': 106, 'offer_image1': 107, 'offer_image3': 108, 'offer_image2': 109, 'offer_image5': 110, 'offer_image4': 111, 'flash_point_unit_of_measure': 112, 'pearl_stringing_method': 113, 'pearl_minimum_color': 114, 'inscription': 115, 'number_of_pearls': 116, 'number_of_stones': 117, 'item_width_unit_of_measure': 118, 'total_diamond_weight': 119, 'item_width': 120, 'item_height': 121, 'stone_weight': 122, 'stone_weight_unit_of_measure': 123, 'size_map': 124, 'size_per_pearl': 125, 'unit_count': 126, 'item_height_unit_of_measure': 127, 'item_diameter_derived': 128, 'item_diameter_unit_of_measure': 129, 'item_length_unit_of_measure': 130, 'item_length': 131, 'total_diamond_weight_unit_of_measure': 132, 'unit_count_type': 133, 'package_length': 134, 'package_weight_unit_of_measure': 135, 'package_height': 136, 'fulfillment_center_id': 137, 'package_length_unit_of_measure': 138, 'package_height_unit_of_measure': 139, 'package_width': 140, 'package_width_unit_of_measure': 141, 'package_weight': 142, 'batteries_required': 143, 'are_batteries_included': 144, 'battery_cell_composition': 145, 'battery_type1': 146, 'battery_type2': 147, 'battery_type3': 148, 'number_of_batteries1': 149, 'number_of_batteries2': 150, 'number_of_batteries3': 151, 'battery_weight': 152, 'battery_weight_unit_of_measure': 153, 'number_of_lithium_metal_cells': 154, 'number_of_lithium_ion_cells': 155, 'lithium_battery_packaging': 156, 'lithium_battery_energy_content': 157, 'lithium_battery_energy_content_unit_of_measure': 158, 'lithium_battery_weight': 159, 'lithium_battery_weight_unit_of_measure': 160, 'supplier_declared_dg_hz_regulation1': 161, 'supplier_declared_dg_hz_regulation2': 162, 'supplier_declared_dg_hz_regulation3': 163, 'supplier_declared_dg_hz_regulation4': 164, 'supplier_declared_dg_hz_regulation5': 165, 'hazmat_united_nations_regulatory_id': 166, 'safety_data_sheet_url': 167, 'item_weight': 168, 'item_weight_unit_of_measure': 169, 'item_volume': 170, 'item_volume_unit_of_measure': 171, 'country_of_origin': 172, 'warranty_type': 173, 'compliance_media_content_type1': 174, 'compliance_media_content_type2': 175, 'compliance_media_content_type3': 176, 'compliance_media_content_type4': 177, 'compliance_media_content_type5': 178, 'compliance_media_content_type6': 179, 'compliance_media_content_type7': 180, 'compliance_media_content_type8': 181, 'compliance_media_content_type9': 182, 'compliance_media_content_type10': 183, 'compliance_media_content_type11': 184, 'compliance_media_content_type12': 185, 'compliance_media_content_type13': 186, 'compliance_media_content_type14': 187, 'compliance_media_content_type15': 188, 'compliance_media_content_type16': 189, 'compliance_media_content_type17': 190, 'flash_point': 191, 'warranty_description': 192, 'external_product_information': 193, 'supplier_declared_material_regulation1': 194, 'supplier_declared_material_regulation2': 195, 'supplier_declared_material_regulation3': 196, 'compliance_media_source_location1': 197, 'compliance_media_source_location2': 198, 'compliance_media_source_location3': 199, 'compliance_media_source_location4': 200, 'compliance_media_source_location5': 201, 'compliance_media_source_location6': 202, 'compliance_media_source_location7': 203, 'compliance_media_source_location8': 204, 'compliance_media_source_location9': 205, 'compliance_media_source_location10': 206, 'compliance_media_source_location11': 207, 'compliance_media_source_location12': 208, 'compliance_media_source_location13': 209, 'compliance_media_source_location14': 210, 'compliance_media_source_location15': 211, 'compliance_media_source_location16': 212, 'compliance_media_source_location17': 213, 'metal_stamp': 214, 'compliance_media_content_language1': 215, 'compliance_media_content_language2': 216, 'compliance_media_content_language3': 217, 'compliance_media_content_language4': 218, 'compliance_media_content_language5': 219, 'compliance_media_content_language6': 220, 'compliance_media_content_language7': 221, 'compliance_media_content_language8': 222, 'compliance_media_content_language9': 223, 'compliance_media_content_language10': 224, 'compliance_media_content_language11': 225, 'compliance_media_content_language12': 226, 'compliance_media_content_language13': 227, 'compliance_media_content_language14': 228, 'compliance_media_content_language15': 229, 'compliance_media_content_language16': 230, 'compliance_media_content_language17': 231, 'ghs_classification_class': 232, 'map_price': 233, 'list_price_with_tax': 234, 'merchant_release_date': 235, 'currency': 236, 'restock_date': 237, 'number_of_items': 238, 'fulfillment_latency': 239, 'offering_can_be_giftwrapped': 240, 'offering_end_date': 241, 'max_order_quantity': 242, 'merchant_shipping_group_name': 243, 'offering_start_date': 244, 'maximum_retail_price': 245, 'product_tax_code': 246, 'offering_can_be_gift_messaged': 247, 'condition_type': 248, 'condition_note': 249, 'sale_price': 250, 'sale_from_date': 251, 'sale_end_date': 252, 'business_price': 253, 'quantity_price_type': 254, 'quantity_price1': 255, 'quantity_lower_bound1': 256, 'quantity_price2': 257, 'quantity_lower_bound2': 258, 'quantity_price3': 259, 'quantity_lower_bound3': 260, 'quantity_price4': 261, 'quantity_lower_bound4': 262, 'quantity_price5': 263, 'quantity_lower_bound5': 264, 'pricing_action': 265, 'unspsc_code': 266, 'national_stock_number': 267}

target_fields_earrings_amz =  [
    'item_sku',
    'item_name',
    'stone_color',
    'style_name',
    'stones_color',
    'item_type_name',
    'occasion_type1',
    'occasion_type2',
    'occasion_type3',
    'occasion_type4',
    'occasion_type5',
    'clasp_type',
    'color_map',
    'material_type_free1',
    'stone_shape',
    'stone_clarity',
    'collection_name',
    'metal_type',
    'material_type',
    'number_of_stones',
    'setting_type',
    'bullet_point1',
    'bullet_point2',
    'bullet_point3',
    'bullet_point4',
    'bullet_point5',
    'bullet_point6',
    'bullet_point7',
    'bullet_point8',
    'bullet_point9',
    'bullet_point10',
    'color_name',
    'stones_shape',
    'theme',
    'back_finding',
    'gem_type',
    'product_description'
]

fixed_values_earrings_meesho = {
    'variation' : 'Free size',
    'GST %': '3',
    'HSN ID' : '71171900',
    'Country of origin' : 'India',
    'Manufacturer Details' : 'Alya jewels',
    'Packer Details' : 'Alya jewels',
    'Brand Name' : 'Alya jewels',
}

target_fields_earrings_meesho =  [
    'Fields + Description:',
    'Product Name',
    'Base Metal',
    'Color',
    'Occasion',
    'Plating',
    'Sizing',
    'Stone Type',
    'Trend',
    'Type',
    'Product Description'
]

# fixed_values_necklace_flipkart , target_fields_necklace_flipkart = get_fixed_values_dict('Necklace_flipkart.xlsx')
# print(fixed_values_necklace_flipkart)
# print()
# print(target_fields_necklace_flipkart)

# fixed_values_necklace_amz , target_fields_necklace_amz = get_fixed_values_dict('Amazon necklace.xlsx')



# print(fixed_values_necklace_amz)
# print()
# print(target_fields_necklace_amz)


fixed_values_necklace_meesho = {
    'variation' : 'Free size',
    'GST %': '3',
    'HSN ID' : '71171900',
    'Country of origin' : 'India',
    'Manufacturer Details' : 'Alya Jewels',
    'Packer Details' : 'Alya Jewels',
    'Net Quantity (N)' : '1',
    'Brand Name' : 'Alya Jewels',
    'Brand' : 'Alya Jewels'
    
}

target_fields_necklace_meesho =  [
    'Fields + Description:',
    'Product Name',
    'Base Metal',
    'Color',
    'Occasion',
    'Plating',
    'Sizing',
    'Stone Type',
    'Trend',
    'Type',
    'Product Description'
]

fixed_values_bracelet_meesho = {
    'variation' : 'Free size',
    'GST %': '3',
    'HSN ID' : '71171900',
    'Country of origin' : 'India',
    'Manufacturer Details' : 'Alya Jewels',
    'Packer Details' : 'Alya Jewels',
    'Net Quantity (N)' : '1',
    'Brand Name' : 'Alya Jewels',
}

target_fields_bracelet_meesho =  [
    'Fields + Description:',
    'Product Name',
    'Base Metal',
    'Closure'
    'Color',
    'Occasion',
    'Plating',
    'Sizing',
    'Stone Type',
    'Trend',
    'Type',
    'Product Description'
]