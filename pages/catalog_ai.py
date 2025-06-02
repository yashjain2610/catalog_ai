import streamlit as st
import os
import base64
import openpyxl
import pandas as pd
import json
import re
import tempfile
import time

import google.generativeai as genai

from langchain.prompts import PromptTemplate
from dotenv import load_dotenv
from PIL import Image
from openpyxl import Workbook

from utils import get_gemini_responses, input_image_setup, input_image_setup_local, encode_image , get_gemini_dims_responses
from prompts import * 
from excel_fields import *

load_dotenv()

genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))


def write_to_excel_meesho(results,filename,target_fields,fixed_values):
    if os.path.exists(filename):
        #print("started loading")
        #print()
        wb = openpyxl.load_workbook(filename)
        #print("finished")
        #print()
        ws = wb.worksheets[1]

        # Read headers dynamically from row 1
        headers = {cell.value: cell.column for cell in ws[3] if cell.value}
        headers = {key.strip().split('\n\n')[0]: value for key, value in headers.items()}

        #print(headers)
        #print()
    else:
        wb = Workbook()
        ws = wb.active  

        # Define default headers
        default_headers = [
            "Image Name", "Type", "Color", "Gemstone", "Pearl Type",
            "Collection", "Occasion", "Piercing Required", "Number of Gemstones",
            "Earring Back Type", "Finish", "Design", "Metal Color", "Diamond Type",
            "Pearl Shape", "Pearl Color", "Search Keywords", "Key Features", "Trend",
            "Closure Type", "Sub Type", "Shape", "Ear Chain", "Earring Set Type",
            "Ornamentation Type", "Trend"
        ]
        
        headers = {name: idx + 1 for idx, name in enumerate(default_headers)}
        ws.append(default_headers)  # Write headers in row 1 if creating a new file

    #print()
    # Combine target fields and fixed value fields
    all_fields = set(target_fields).union(fixed_values.keys())

    # Get column indices for all relevant fields
    target_columns = {field: headers[field] for field in all_fields if field in headers}
    # print(target_columns)

    # **Write output from row 6 onwards**
    row_idx = 1
    while ws.cell(row=row_idx, column=1).value:  # Assuming column 1 is "Image Name" or always filled
        row_idx += 1

    for image_name, response , description in results:
        #print(response)
        # answers = response.split("\n")  
        # answers = [ans.strip() for ans in answers]
        # answers.insert(0, image_name)

        # print()
        # print(answers)  
        # print()

        # Create a dictionary mapping field names to values
        # print(target_fields)
        # field_values = {field: answers[i] if i < len(answers) else "" for i, field in enumerate(target_fields)}
        # print()
        # print(field_values)
        field_values = response

        # Write values only in the specified fields
        for key, val in fixed_values.items():
            field_values[key] = val

        field_values["Product Description"] = description
        field_values["Fields + Description:"] = os.path.splitext(image_name)[0]

        for field, col_idx in target_columns.items():
            ws.cell(row=row_idx, column=col_idx, value=field_values.get(field, ""))

        row_idx += 1  # Move to the next row

    wb.save(filename)


def write_to_excel_flipkart(results,filename,target_fields,fixed_values):
    if os.path.exists(filename):
        #print("started loading")
        #print()
        wb = openpyxl.load_workbook(filename)
        #print("finished")
        #print()
        ws = wb.worksheets[2]

        # Read headers dynamically from row 1
        headers = {cell.value: cell.column for cell in ws[1] if cell.value}
        #print(headers)
        #print()
    else:
        wb = Workbook()
        ws = wb.active  

        # Define default headers
        default_headers = [
            "Image Name", "Type", "Color", "Gemstone", "Pearl Type",
            "Collection", "Occasion", "Piercing Required", "Number of Gemstones",
            "Earring Back Type", "Finish", "Design", "Metal Color", "Diamond Type",
            "Pearl Shape", "Pearl Color", "Search Keywords", "Key Features", "Trend",
            "Closure Type", "Sub Type", "Shape", "Ear Chain", "Earring Set Type",
            "Ornamentation Type", "Trend"
        ]
        
        headers = {name: idx + 1 for idx, name in enumerate(default_headers)}
        ws.append(default_headers)  # Write headers in row 1 if creating a new file

    #print()
    # Combine target fields and fixed value fields
    all_fields = set(target_fields).union(fixed_values.keys())

    # Get column indices for all relevant fields
    target_columns = {field: headers[field] for field in all_fields if field in headers}
    # print(target_columns)

    # **Write output from row 6 onwards**
    row_idx = 1
    while ws.cell(row=row_idx, column=7).value:  # Assuming column 1 is "Image Name" or always filled
        row_idx += 1
    #print(row_idx)  
    for image_name, response , description in results:
        field_values = response

        # Write values only in the specified fields
        for key, val in fixed_values.items():
            field_values[key] = val

        field_values["Description"] = description
        field_values["Seller SKU ID"] = os.path.splitext(image_name)[0]

        for field, col_idx in target_columns.items():
            ws.cell(row=row_idx, column=col_idx, value=field_values.get(field, ""))

        row_idx += 1  # Move to the next row

    wb.save(filename)


def write_to_excel_amz_xl(results,filename,target_fields,fixed_values):
    if os.path.exists(filename):
        print("started loading")
        print()
        wb = openpyxl.load_workbook(filename)
        print("finished")
        print()
        ws = wb.worksheets[0]

        # Read headers dynamically from row 1
        headers = {cell.value: cell.column for cell in ws[3] if cell.value}
        print(headers)
        print()
    else:
        wb = Workbook()
        ws = wb.active  

        # Define default headers
        default_headers = [
            "Image Name", "Type", "Color", "Gemstone", "Pearl Type",
            "Collection", "Occasion", "Piercing Required", "Number of Gemstones",
            "Earring Back Type", "Finish", "Design", "Metal Color", "Diamond Type",
            "Pearl Shape", "Pearl Color", "Search Keywords", "Key Features", "Trend",
            "Closure Type", "Sub Type", "Shape", "Ear Chain", "Earring Set Type",
            "Ornamentation Type", "Trend"
        ]
        
        headers = {name: idx + 1 for idx, name in enumerate(default_headers)}
        ws.append(default_headers)  # Write headers in row 1 if creating a new file

    #print()
    # Combine target fields and fixed value fields
    all_fields = set(target_fields).union(fixed_values.keys())

    # Get column indices for all relevant fields
    target_columns = {field: headers[field] for field in all_fields if field in headers}
    #print(target_columns)

    # **Write output from row 6 onwards**
    row_idx = 1
    while ws.cell(row=row_idx, column=2).value:  # Assuming column 1 is "Image Name" or always filled
        row_idx += 1
    #print(row_idx)  
    for image_name, response , description in results:
        # print(response)
        # answers = response.split("\n")  
        # answers = [ans.strip() for ans in answers]
        # answers.insert(0, os.path.splitext(image_name)[0])

        # print()
        # print(answers)  
        # print()

        # Create a dictionary mapping field names to values
        # print(target_fields)
        # field_values = {field: answers[i] if i < len(answers) else "" for i, field in enumerate(target_fields)}
        # print()
        # print(field_values)
        field_values = response

        # Write values only in the specified fields
        for key, val in fixed_values.items():
            field_values[key] = val

        field_values["product_description"] = description
        field_values["item_sku"] = os.path.splitext(image_name)[0]

        for field, col_idx in target_columns.items():
            ws.cell(row=row_idx, column=col_idx, value=field_values.get(field, ""))

        row_idx += 1  # Move to the next row

    wb.save(filename)


# def save_to_excel_flipkart():
#         uploaded_file = st.file_uploader("📤 Upload your Excel file", type=["xlsx"] , key="up2")

#         if uploaded_file:
#             st.info("Now click below to save the results to your uploaded Excel.")
#             if st.button("💾 Save Results to Excel", key="bu1"):
#                 with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
#                     tmp.write(uploaded_file.read())
#                     temp_filename = tmp.name

#                 # Save to Excel
#                 write_to_excel_flipkart(
#                     st.session_state["results_ready_flp"],
#                     temp_filename,
#                     target_fields_earrings,
#                     fixed_values_earrings
#                 )

#                 # Download button
#                 with open(temp_filename, "rb") as f:
#                     st.download_button(
#                         label="📥 Download Updated Excel",
#                         data=f,
#                         file_name="Updated_Earrings_Flipkart.xlsx",
#                         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                         key="dbu1"
#                     )




def on_click_uploaded(uploaded_files, input_prompt_list,format,dimensions_prompt_list=[]):
        #results = []
        formated_string = f"edited_results_{format}"
        if formated_string not in st.session_state:
            st.session_state[formated_string] = []
        if not uploaded_files:
            st.error("No images uploaded!")
            return

        count = 0

        for uploaded_file in uploaded_files:
            image = Image.open(uploaded_file)

            st.image(
                image,
                caption=f"Processing: {uploaded_file.name}",
                use_container_width=True
            )

            data_key = f"{uploaded_file.name}_data_{format}"
            desc_key = f"{uploaded_file.name}_desc_{format}"
            initialized_key = f"{uploaded_file.name}_initialized_{format}"

        # Only call AI APIs and set data once
            if not st.session_state.get(initialized_key):
                image_data = input_image_setup(uploaded_file)

                response_list = get_gemini_responses(
                    "Analyze this image carefully.",
                    image_data,
                    input_prompt_list
                )

                dimensions_response_dict = {}
                if dimensions_prompt_list:
                    dimensions_response = get_gemini_dims_responses(
                        "Analyze this image carefully.",
                        image_data,
                        dimensions_prompt_list
                    )

                    cleaned_dims = re.sub(r"```(?:json)?\s*([\s\S]*?)\s*```", r"\1", dimensions_response.strip(), flags=re.IGNORECASE)
                    dimensions_response_dict = json.loads(cleaned_dims.strip())

                description_response = response_list[0] if len(response_list) > 0 else "No description response."
                question_response = response_list[1] if len(response_list) > 1 else "No answer response."

                cleaned = re.sub(r"```(?:json)?\s*([\s\S]*?)\s*```", r"\1", question_response.strip(), flags=re.IGNORECASE)

                question_response_dict = json.loads(cleaned.strip())
                if dimensions_response_dict:
                    question_response_dict.update(dimensions_response_dict)

            # Store in session state
                st.session_state[data_key] = question_response_dict
                st.session_state[desc_key] = description_response
                st.session_state[initialized_key] = True

        # Get from session state
            question_response_dict = st.session_state[data_key]
            description_response = st.session_state[desc_key]

            st.subheader(f"📝 Description for {uploaded_file.name}:")
            st.write(description_response)

            st.subheader(f"🔍 Response for {uploaded_file.name}:")
            st.write(question_response_dict)

            st.session_state[formated_string].append((uploaded_file.name, question_response_dict, description_response))

            count += 1
            if count%5 == 0:
                time.sleep(45)

        return st.session_state[formated_string]

def on_click(image_folder_name, input_prompt_list):
    #results = []
    if "edited_results" not in st.session_state:
         st.session_state["edited_results"] = []
    image_folder = image_folder_name  # Path to the folder containing images
    image_files = [f for f in os.listdir(image_folder) if f.endswith((".JPG", ".jpeg", ".png", "jpg"))]

    if not image_files:
        st.error("No images found in the folder!")
        return

    for image_file in image_files:
        image_path = os.path.join(image_folder, image_file)
        image = Image.open(image_path)

        st.image(
            image,
            caption=f"Processing: {image_file}",
            use_container_width=True
        )

        # Unique keys for session state
        data_key = f"{image_file}_data"
        desc_key = f"{image_file}_desc"
        initialized_key = f"{image_file}_initialized"

        # Only call AI APIs and set data once
        if not st.session_state.get(initialized_key):
            image_data = input_image_setup_local(image_path)

            response_list = get_gemini_responses(
                "Analyze this image carefully.",
                image_data,
                input_prompt_list
            )

            description_response = response_list[0] if len(response_list) > 0 else "No description response."
            question_response = response_list[1] if len(response_list) > 1 else "No answer response."

            cleaned = re.sub(r"```(?:json)?\s*([\s\S]*?)\s*```", r"\1", question_response.strip(), flags=re.IGNORECASE)
            # cleaned_dims = re.sub(r"```(?:json)?\s*([\s\S]*?)\s*```", r"\1", dimensions_response.strip(), flags=re.IGNORECASE)

            question_response_dict = json.loads(cleaned.strip())
            # dimensions_response_dict = json.loads(cleaned_dims.strip())
            # question_response_dict.update(dimensions_response_dict)

            # Store in session state
            st.session_state[data_key] = question_response_dict
            st.session_state[desc_key] = description_response
            st.session_state[initialized_key] = True

        # Get from session state
        question_response_dict = st.session_state[data_key]
        description_response = st.session_state[desc_key]

        st.subheader(f"📝 Description for {image_file}:")
        st.write(description_response)

        st.subheader(f"🔍 Response for {image_file}:")
        st.write(question_response_dict)

        # st.subheader(f"📏 Dimensions for {image_file}:")
        # st.write(dimensions_response)

        st.session_state["edited_results"].append((image_file, question_response_dict, description_response))

    return st.session_state["edited_results"]



def main():
    st.set_page_config(page_title="Jewelry Catalog Assistant")
    st.title("Jewelry Catalog Assistant")

    ## logic for uploading files
    # File uploader for image
    uploaded_files = st.file_uploader("Choose an image...", type=["jpg", "jpeg", "png", "webp"] , accept_multiple_files=True , key="up1")
    image = ""

    ### uploading logic

    if uploaded_files is not None:
        for uploaded_file in uploaded_files:
            image = Image.open(uploaded_file)
            #image_base64 = encode_image(uploaded_file)
            st.image(image, caption="Uploaded Image.", use_container_width=True)
    
    ##logic for generating content from local images

    ### excel files uploading
    st.header("upload excel files for output here")
    excel_file_flp_ear = st.file_uploader("📤 Upload your Excel file of flipkart for earrings", type=["xlsx"] , key="up2")
    temp_filename_flp_ear = None
    if excel_file_flp_ear:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(excel_file_flp_ear.read())
            temp_filename_flp_ear = tmp.name

    excel_file_amz_ear = st.file_uploader("📤 Upload your Excel file of amazon for earrings", type=["xlsx"] , key="up3")
    temp_filename_amz_ear = None
    if excel_file_amz_ear:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(excel_file_amz_ear.read())
            temp_filename_amz_ear = tmp.name

    excel_file_mee_ear = st.file_uploader("📤 Upload your Excel file of meesho for earrings", type=["xlsx"] , key="up4")
    temp_filename_mee_ear = None
    if excel_file_mee_ear:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(excel_file_mee_ear.read())
            temp_filename_mee_ear = tmp.name

    excel_file_flp_neck = st.file_uploader("📤 Upload your Excel file of flipkart for necklaces", type=["xlsx"] , key="up5")
    temp_filename_flp_neck = None
    if excel_file_flp_neck:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(excel_file_flp_neck.read())
            temp_filename_flp_neck = tmp.name

    excel_file_amz_neck = st.file_uploader("📤 Upload your Excel file of amazon for necklaces", type=["xlsx"] , key="up6")
    temp_filename_amz_neck = None
    if excel_file_amz_neck:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(excel_file_amz_neck.read())
            temp_filename_amz_neck = tmp.name

    excel_file_mee_neck = st.file_uploader("📤 Upload your Excel file of meesho for necklaces", type=["xlsx"] , key="up7")
    temp_filename_mee_neck = None
    if excel_file_mee_neck:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(excel_file_flp_neck.read())
            temp_filename_mee_neck = tmp.name
    
    excel_file_flp_brc = st.file_uploader("📤 Upload your Excel file of flipkart for bracelet", type=["xlsx"] , key="up8")
    temp_filename_flp_brc = None
    if excel_file_flp_brc:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(excel_file_flp_brc.read())
            temp_filename_flp_brc = tmp.name
    
    excel_file_mee_brc = st.file_uploader("📤 Upload your Excel file of meesho for bracelet", type=["xlsx"] , key="up9")
    temp_filename_mee_brc = None
    if excel_file_mee_brc:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(excel_file_mee_brc.read())
            temp_filename_mee_brc = tmp.name

    
        
    

    ### EARRINGS

    st.header("Earrings")
    
    ## EARRINGS flipkart

    # submit_local = st.button("process local images for earrings(flipkart format)")
    submit_uploaded_earrings = st.button("process uploaded images for earrings(flipkart format)")
    input_prompt_earrings = [prompt_description_earrings_flipkart, prompt_questions_earrings_flipkart]
    dimensions_prompt_list = [prompt_dimensions_earrings_flipkart]

    if submit_uploaded_earrings:
        results = on_click_uploaded(uploaded_files,input_prompt_earrings,"flp",dimensions_prompt_list)
        st.session_state["results_ready_flp_ear"] = results
        st.success("✅ Results generated. Please upload Excel file to save them.")

    if "results_ready_flp_ear" in st.session_state:
        if excel_file_flp_ear:
            st.info("Now click below to save the results to your uploaded Excel.")
            if st.button("💾 Save Results to Excel", key="bu1"):
                # Save to Excel
                write_to_excel_flipkart(
                    st.session_state["results_ready_flp_ear"],
                    temp_filename_flp_ear,
                    target_fields_earrings,
                    fixed_values_earrings
                )
        #write_to_excel_flipkart(results,"Earrings.xlsx",target_fields_earrings,fixed_values_earrings)
        #st.success(f"✅ Results have been saved to earrings.xlsx")



    # if submit_local:
    #     results = on_click(
    #         image_folder_name="earrings images",
    #         input_prompt_list=input_prompt_earrings,
    #         dimesions_prompt_list=dimensions_prompt_list
    #     )
    #     st.session_state["results_ready"] = results
    #     st.success("✅ Results generated. Please upload Excel file to save them.")

    # if "results_ready" in st.session_state:
    #     uploaded_file = st.file_uploader("📤 Upload your Excel file", type=["xlsx"])

    #     if uploaded_file:
    #         st.info("Now click below to save the results to your uploaded Excel.")
    #         if st.button("💾 Save Results to Excel"):
    #             with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
    #                 tmp.write(uploaded_file.read())
    #                 temp_filename = tmp.name

    #         # Save to Excel
    #             write_to_excel_flipkart(
    #                 st.session_state["results_ready"],
    #                 temp_filename,
    #                 target_fields_earrings,
    #                 fixed_values_earrings
    #             )

    #         # Download button
    #             with open(temp_filename, "rb") as f:
    #                 st.download_button(
    #                     label="📥 Download Updated Excel",
    #                     data=f,
    #                     file_name="Updated_Earrings_Flipkart.xlsx",
    #                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    #                 )

        # write_to_excel_flipkart(results,"Earrings_flipkart.xlsx",target_fields_earrings,fixed_values_earrings)
        # st.success(f"✅ Results have been saved to earrings.xlsx")  

    ###  EARRINGS AMAZON

    input_prompt_earrings_amazon = [prompt_description_earrings_amz,prompt_questions_earrings_amz]
    dimensions_prompt_list_amz = [prompt_dimensions_earrings_amz]

    # submit_local_earrings_amazon = st.button("process local images for earrings(amazon format)")
    submit_uploaded_earrings_amazon = st.button("process uploaded images for earrings(amazon format)")

    if submit_uploaded_earrings_amazon:
        results = on_click_uploaded(uploaded_files,input_prompt_earrings_amazon,"amz", dimensions_prompt_list=dimensions_prompt_list_amz)
        st.session_state["results_ready_amz_ear"] = results
        st.success("✅ Results generated. Please upload Excel file to save them.")

    if "results_ready_amz_ear" in st.session_state:
        if excel_file_amz_ear:
            st.info("Now click below to save the results to your uploaded Excel.")
            if st.button("💾 Save Results to Excel", key="bu2"):
                # Save to Excel
                write_to_excel_amz_xl(
                    st.session_state["results_ready_amz_ear"],
                    temp_filename_amz_ear,
                    target_fields_earrings_amz,
                    fixed_values_earrings_amz
                )

    # if submit_local_earrings_amazon:
    #     results = on_click("earrings images",input_prompt_earrings_amazon)
    #     write_to_excel_amz_xl(results,filename="earrings_amz.xlsx",target_fields=target_fields_earrings_amz,fixed_values=fixed_values_earrings_amz)
    #     st.success(f"✅ Results have been saved to Earrings_amz.xlsx")

    ### Earrings Meesho

    input_prompt_earrings_meesho = [prompt_description_earrings_meesho,prompt_questions_earrings_meesho]

    # submit_local_earrings_meesho = st.button("process local images for earrings(meesho format)")
    submit_uploaded_earrings_meesho = st.button("process uploaded images for eaarings (meesho format)")

    if submit_uploaded_earrings_meesho:
        results = on_click_uploaded(uploaded_files,input_prompt_earrings_meesho, "mee")
        st.session_state["results_ready_mee_ear"] = results
        st.success("✅ Results generated")

    if "results_ready_mee_ear" in st.session_state:
        if excel_file_mee_ear:
            st.info("Now click below to save the results to your uploaded Excel.")
            if st.button("💾 Save Results to Excel", key = "bu3"):
                # Save to Excel
                write_to_excel_meesho(
                    st.session_state["results_ready_mee_ear"],
                    temp_filename_mee_ear,
                    target_fields_earrings_meesho,
                    fixed_values_earrings_meesho
                )

        # write_to_excel_meesho(results,filename="Earrings_meesho.xlsx",target_fields=target_fields_earrings_meesho,fixed_values=fixed_values_earrings_meesho)
        # st.success(f"✅ Results have been saved to Earrings_Meesho.xlsx")

    # if submit_local_earrings_meesho:
    #     results = on_click("earrings images",input_prompt_earrings_meesho)
    #     write_to_excel_meesho(results,filename="Earrings_meesho.xlsx",target_fields=target_fields_earrings_meesho,fixed_values=fixed_values_earrings_meesho)
    #     st.success(f"✅ Results have been saved to Earrings_Meesho.xlsx")


### NECKLACE

    st.header("Necklace")


##NECKLACE(flipkart)

    results_necklace = []

    input_prompt_necklace = [prompt_description_necklace_flipkart, prompt_questions_necklace_flipkart]
    dimensions_prompt_list_necklace = [prompt_dimensions_necklace_flipkart]

    #submit_local_necklace = st.button("process local images for necklace (flipkart format)")
    submit_uploaded_necklace = st.button("process uploaded images for necklace (flipkart format)")
    
    if submit_uploaded_necklace:
        results_necklace = on_click_uploaded(uploaded_files,input_prompt_necklace,"flp",dimensions_prompt_list=dimensions_prompt_list_necklace)
        st.session_state["results_ready_flp_neck"] = results_necklace
        st.success("✅ Results generated.")

    if "results_ready_flp_neck" in st.session_state:
        if excel_file_flp_neck:
            st.info("Now click below to save the results to your uploaded Excel.")
            if st.button("💾 Save Results to Excel", key="bu4"):
                # Save to Excel
                write_to_excel_flipkart(
                    st.session_state["results_ready_flp_neck"],
                    temp_filename_flp_neck,
                    target_fields_necklace_flipkart,
                    fixed_values_necklace_flipkart
                )



        # print(results_necklace)

        # write_to_excel_flipkart(results_necklace,filename="Necklace_flipkart.xlsx",target_fields=target_fields_necklace_flipkart,fixed_values=fixed_values_necklace_flipkart)
        # st.success("✅ Results have been saved to necklace.xlsx")

    # if submit_local_necklace:

    #     results_necklace = on_click("necklace images",input_prompt_necklace)

    #     write_to_excel_flipkart(results_necklace,filename="Necklace_flipkart.xlsx",target_fields=target_fields_necklace_flipkart,fixed_values=fixed_values_necklace_flipkart)
    #     st.success("✅ Results have been saved to necklace.xlsx")

### Necklace Amazon

    input_prompt_necklace_amazon = [prompt_description_necklace_amz, prompt_questions_necklace_amz]

    submit_uploaded_necklace_amazon = st.button("process uploaded images for necklace(amazon format)")
    # submit_local_necklace_amazon = st.button("process local images for necklace(amazon format)")

    if submit_uploaded_necklace_amazon:
        results_necklace = on_click_uploaded(uploaded_files,input_prompt_necklace_amazon,"amz")
        st.session_state["results_ready_amz_neck"] = results_necklace
        st.success("✅ Results generated.")
    if "results_ready_amz_neck" in st.session_state:
        if excel_file_amz_neck:
            st.info("Now click below to save the results to your uploaded Excel.")
            if st.button("💾 Save Results to Excel", key="bu5"):
                # Save to Excel
                write_to_excel_amz_xl(
                    st.session_state["results_ready_amz_neck"],
                    temp_filename_amz_neck,
                    target_fields_necklace_amz,
                    fixed_values_necklace_amz
                )



    # if submit_local_necklace_amazon:
    #     results = on_click("necklace images", input_prompt_necklace_amazon)
    #     # save_to_excel = st.button("save output to excel")
    #     # if save_to_excel:
    #     #     print(results)
    #     write_to_excel_amz_xl(results,filename="Necklace amz.xlsx", target_fields= target_fields_necklace_amz,fixed_values=fixed_values_necklace_amz)
    #     st.success(f"✅ Results have been saved to Neclace amz.xlsx")



### Necklace Meesho
    input_prompt_necklace_meesho = [prompt_description_necklace_meesho,prompt_questions_necklace_meesho]
    # submit_local_necklace_meesho = st.button("process local images for necklace(meesho format)")
    submit_uploaded_necklace_meesho = st.button("process uploaded images for necklace (meesho format)")

    if submit_uploaded_necklace_meesho:
        results_necklace = on_click_uploaded(uploaded_files,input_prompt_necklace_meesho,"mee")
        st.session_state["results_ready_mee_neck"] = results_necklace
        st.success("✅ Results generated.")
    if "results_ready_mee_neck" in st.session_state:
        if excel_file_mee_neck:
            st.info("Now click below to save the results to your uploaded Excel.")
            if st.button("💾 Save Results to Excel", key="bu6"):
                # Save to Excel
                write_to_excel_meesho(
                    st.session_state["results_ready_mee_neck"],
                    temp_filename_mee_neck,
                    target_fields_necklace_meesho,
                    fixed_values_necklace_meesho
                )

    # if submit_local_necklace_meesho:
    #     results = on_click("necklace images",input_prompt_necklace_meesho)

    #     write_to_excel_meesho(results,filename="Necklace_meesho.xlsx",target_fields=target_fields_necklace_meesho,fixed_values=fixed_values_necklace_meesho)
    #     st.success(f"✅ Results have been saved to Necklace_Meesho.xlsx")

### bracelet

    st.header("Bracelet")
    

### BRACELET (flipkart)

    input_prompt_bracelet = [prompt_description_bracelet_flipkart,prompt_questions_bracelet_flipkart]
    dimensions_prompt_list_bracelet = [prompt_dimensions_bracelet_flipkart]

    #submit_local_bracelet = st.button("process local images for bracelet (flipkart format)")
    submit_uploaded_bracelet = st.button("process uploaded images for bracelet(flipkart format)")

    if submit_uploaded_bracelet:
        results_bracelet = on_click_uploaded(uploaded_files,input_prompt_bracelet,"flp",dimensions_prompt_list=dimensions_prompt_list_bracelet)
        st.session_state["results_ready_flp_brc"] = results_bracelet
        st.success("✅ Results generated.")

    if "results_ready_flp_brc" in st.session_state:
        if excel_file_flp_brc:
            st.info("Now click below to save the results to your uploaded Excel.")
            if st.button("💾 Save Results to Excel", key="bu7"):
                # Save to Excel
                write_to_excel_flipkart(
                    st.session_state["results_ready_flp_brc"],
                    temp_filename_flp_brc,
                    target_fields_bracelet_flipkart,
                    fixed_values_bracelet_flipkart
                )

    # if submit_local_bracelet:
    #     results = on_click("bracelet images",input_prompt_bracelet)
    #     write_to_excel_flipkart(results,"bracelet_flipkart.xlsx",target_fields_earrings,fixed_values_earrings)
    #     st.success(f"✅ Results have been saved to braclet excel")

### bracelet meesho

    input_prompt_bracelet_meesho = [prompt_description_bracelet_meesho,prompt_questions_bracelet_meesho]
    # submit_local_bracelet_meesho = st.button("process local images for bracelet(meesho format)")

    submit_uploaded_bracelet_meesho = st.button("process uploaded images for bracelet (meesho format)")

    if submit_uploaded_bracelet_meesho:
        results_bracelet = on_click_uploaded(uploaded_files,input_prompt_bracelet_meesho,"mee")
        st.session_state["results_ready_mee_brc"] = results_bracelet
        st.success("✅ Results generated.")
    if "results_ready_mee_brc" in st.session_state:
        if excel_file_mee_brc:
            st.info("Now click below to save the results to your uploaded Excel.")
            if st.button("💾 Save Results to Excel", key="bu8"):
                # Save to Excel
                write_to_excel_meesho(
                    st.session_state["results_ready_mee_brc"],
                    temp_filename_mee_brc,
                    target_fields_bracelet_meesho,
                    fixed_values_bracelet_meesho
                )


    # if submit_local_bracelet_meesho:
    #     results = on_click("bracelet images",input_prompt_bracelet_meesho)
    #     write_to_excel_meesho(results,filename="bracelet_meesho.xlsx",target_fields=target_fields_bracelet_meesho,fixed_values=fixed_values_bracelet_meesho)
    #     st.success(f"✅ Results have been saved to bracelet_Meesho.xlsx")


### DOWNLOAD BUTTONS
    st.header("Download Excel")
    if excel_file_flp_ear:
        # Download button
        with open(temp_filename_flp_ear, "rb") as f:
                st.download_button(
                        label="📥 Download Updated Excel for earrings(flipkart)",
                        data=f,
                        file_name="Updated_Earrings_Flipkart.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dbu1"
                )
    if excel_file_amz_ear:
        # Download button
        with open(temp_filename_amz_ear, "rb") as f:
            st.download_button(
            label="📥 Download Updated Excel for earrings (amazon)",
            data=f,
            file_name="Updated_Earrings_amz.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )      
    if excel_file_mee_ear:
        with open(temp_filename_mee_ear, "rb") as f:
            st.download_button(
            label="📥 Download Updated Excel for earrings (meesho)",
            data=f,
            file_name="Updated_Earrings_meesho.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    if excel_file_flp_neck:
        with open(temp_filename_flp_neck, "rb") as f:
            st.download_button(
            label="📥 Download Updated Excel for necklace (flipkart)",
            data=f,
            file_name="Updated_necklace_flipkart.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    if excel_file_amz_neck:
        with open(temp_filename_amz_neck, "rb") as f:
            st.download_button(
            label="📥 Download Updated Excel for necklace (amazon)",
            data=f,
            file_name="Updated_necklace_amazon.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    if excel_file_mee_neck:
        with open(temp_filename_mee_neck, "rb") as f:
            st.download_button(
            label="📥 Download Updated Excel for necklace (meesho)",
            data=f,
            file_name="Updated_necklace_meesho.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    if excel_file_flp_brc:
        with open(temp_filename_flp_brc, "rb") as f:
            st.download_button(
            label="📥 Download Updated Excel for bracelet (flipkart)",
            data=f,
            file_name="Updated_bracelet_flipkart.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    if excel_file_mee_brc:
        with open(temp_filename_mee_brc, "rb") as f:
            st.download_button(
            label="📥 Download Updated Excel for bracelet (meesho)",
            data=f,
            file_name="Updated_necklace_meesho.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

     
        


if __name__ == "__main__":
    main()


##stone color
